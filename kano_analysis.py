from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws.title = "Kano Analysis"

WHITE      = "FFFFFF"
DARK_HDR   = "1F3864"
LIGHT_GRAY = "F2F2F2"

# Action text colors (colored text like Mercadona example)
ACTION_FONT = {
    "Keep":        "000000",
    "Iterate":     "C55A11",
    "Investigate": "1F4E79",
    "Add to test": "7030A0",
    "Cut":         "C00000",
}

def make_border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def dashed_right():
    return Border(right=Side(style="dashed", color="4472C4"))

# ── Title ──────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
ws["A1"] = "My Match Olympics — Kano Feature Analysis"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
ws["A2"] = "Fill in your features below. The chart updates automatically."
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 16

# ── Column headers ─────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 34
header_data = [
    ("A", "#"),
    ("B", "Feature"),
    ("C", "Kano Category"),
    ("D", "Importance\n(1-5)"),
    ("E", "Confidence (1-5)"),
    ("F", "Action"),
    ("G", "Notes / Assumption to test"),
]
for col, label in header_data:
    c = ws[f"{col}4"]
    c.value = label
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=DARK_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = make_border()

# ── Feature rows ───────────────────────────────────────────────────────────────
rows = [
    (1,  "Personality quiz → athlete match",
         "Performer",   5, 4, "Keep",
         "Core value prop. More personalisation = more emotional connection with athlete."),
    (2,  "Business sponsorship metrics dashboard",
         "Must-have",   5, 4, "Keep",
         "Without this, B2B value is zero. The metrics dashboard IS the product for sponsors."),
    (3,  "Follow athlete + persistent news feed",
         "Performer",   4, 5, "Keep",
         "Probably linear with engagement. Need to test if daily news or weekly digest is enough."),
    (4,  "Athlete self-service onboarding portal",
         "Must-have",   4, 2, "Investigate",
         "Hypothesis: must-have for data quality, but athlete adoption unproven. Test with 5 real athletes."),
    (5,  "Archetype reveal animation (slam effect)",
         "Delighter",   3, 5, "Keep",
         "Pure delight assumption. Already built — high emotional impact, shareable moment."),
    (6,  "Medal count & credentials display",
         "Delighter",   3, 4, "Keep",
         "Could shift to performer for B2B users. Credential proof drives sponsor trust."),
    (7,  "Deal value estimates per athlete",
         "Performer",   4, 3, "Iterate",
         "Linear in usefulness. Worth refining once athlete self-reported pricing feeds in."),
    (8,  "Brand safety score",
         "Performer",   4, 2, "Investigate",
         "High B2B value but model-generated — may mislead buyers. Needs validation with real sponsors."),
    (9,  "Share quiz result card on social",
         "Delighter",   2, 2, "Add to test",
         "Pure assumption. Might be a delighter, might be indifferent. Easy and cheap to A/B test."),
    (10, "Export shortlist to CSV",
         "Performer",   3, 3, "Iterate",
         "Linear for power users. Business teams need this for internal reporting workflows."),
    (11, "Push notifications for athlete news",
         "Reverse",     2, 4, "Cut",
         "Likely reverse — adds noise. Web app not native. Users have enough notifications already."),
    (12, "Compare two athletes side-by-side",
         "Indifferent",  2, 2, "Cut",
         "Suspect indifferent for most users. Niche B2B use case, not core to value prop."),
    (13, "Athlete pricing visible to fans",
         "Indifferent",  1, 3, "Cut",
         "Fans don't care about deal prices. Showing it harms athlete trust. Business-only data."),
]

for i, (num, feat, kano, imp, conf, action, notes) in enumerate(rows):
    r = i + 5
    ws.row_dimensions[r].height = 40
    bg = WHITE if i % 2 == 0 else LIGHT_GRAY

    vals = [num, feat, kano, imp, conf, action, notes]
    for col_letter, val in zip(["A","B","C","D","E","F","G"], vals):
        c = ws[f"{col_letter}{r}"]
        c.value  = val
        c.fill   = PatternFill("solid", fgColor=bg)
        c.border = make_border()
        c.alignment = Alignment(
            vertical="center", wrap_text=True,
            horizontal="center" if col_letter in ("A","D","E") else "left"
        )
        c.font = Font(name="Calibri", size=10)

        # Action: bold colored text, no background
        if col_letter == "F":
            c.font = Font(name="Calibri", size=10, bold=True,
                          color=ACTION_FONT.get(action, "000000"))
            c.alignment = Alignment(horizontal="center", vertical="center")

# ── Column widths ──────────────────────────────────────────────────────────────
for col, w in [("A",5),("B",36),("C",14),("D",13),("E",14),("F",12),("G",54)]:
    ws.column_dimensions[col].width = w

# dashed separator after col G
for r in range(4, 4 + len(rows) + 1):
    c = ws[f"G{r}"]
    c.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="dashed", color="4472C4"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

# ── Summary table ──────────────────────────────────────────────────────────────
kano_counts = {}
for _, _, kano, *_ in rows:
    kano_counts[kano] = kano_counts.get(kano, 0) + 1

# header
for col, label in [("I","Category"), ("J","Count")]:
    c = ws[f"{col}4"]
    c.value = label
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=DARK_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = make_border()

cat_order = ["Must-have", "Performer", "Delighter", "Indifferent", "Reverse"]
for i, cat in enumerate(cat_order):
    r = i + 5
    ws[f"I{r}"].value = cat
    ws[f"J{r}"].value = kano_counts.get(cat, 0)
    bg = WHITE if i % 2 == 0 else LIGHT_GRAY
    for col in ("I","J"):
        c = ws[f"{col}{r}"]
        c.font      = Font(name="Calibri", size=10, bold=(col=="J"))
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()

ws.column_dimensions["I"].width = 14
ws.column_dimensions["J"].width = 8

# ── Bar chart ──────────────────────────────────────────────────────────────────
chart = BarChart()
chart.type   = "bar"
chart.title  = "Features by Kano Category"
chart.y_axis.title = None
chart.x_axis.title = "Count"
chart.style  = 10
chart.width  = 14
chart.height = 10
chart.legend = None

chart.add_data(Reference(ws, min_col=10, min_row=4, max_row=9), titles_from_data=True)
chart.set_categories(Reference(ws, min_col=9, min_row=5, max_row=9))
ws.add_chart(chart, "I11")

ws.freeze_panes = "A5"

out = r"C:\Users\Jad Zoghaib\OneDrive\Desktop\my_match_olympics\kano_analysis.xlsx"
wb.save(out)
print(f"Saved: {out}")
